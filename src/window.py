import os
import pickle
from pathlib import Path

from PySide6.QtCore import Qt
from PySide6.QtWidgets import (
    QMainWindow,
    QWidget,
    QPushButton,
    QVBoxLayout,
    QHBoxLayout,
    QFileDialog,
    QListWidget,
    QLabel,
    QListWidgetItem,
    QMessageBox,
    QInputDialog,
    QRadioButton,
)
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

from .image_displayer import ImageDisplayer, PDFFile
from .ocr_tools import perform_ocr
from .rects import Rect, PickleRect


class MainWindow(QMainWindow):
    def __init__(self):
        super().__init__()
        self.setWindowTitle("Screenshot maker")

        self.central_widget = QWidget()
        self.setCentralWidget(self.central_widget)
        self.central_layout = QHBoxLayout()
        self.central_widget.setLayout(self.central_layout)

        self.left_layout = QVBoxLayout()
        self.central_layout.addLayout(self.left_layout)

        self.load_pdf_button = QPushButton("Load PDF")
        self.left_layout.addWidget(self.load_pdf_button)

        self.pdf_buttons_layout = QHBoxLayout()
        self.left_layout.addLayout(self.pdf_buttons_layout)

        self.loaded_pdfs_label = QLabel("Loaded PDFs:")
        self.pdf_buttons_layout.addWidget(self.loaded_pdfs_label)

        self.delete_selected_pdf_button = QPushButton("Delete selected PDF")
        self.delete_selected_pdf_button.clicked.connect(self.delete_selected_pdf)
        self.pdf_buttons_layout.addWidget(self.delete_selected_pdf_button)

        self.delete_all_pdfs_button = QPushButton("Delete all PDFs")
        self.delete_all_pdfs_button.clicked.connect(self.delete_all_pdfs)
        self.pdf_buttons_layout.addWidget(self.delete_all_pdfs_button)

        self.files_listwidget = QListWidget()
        self.files_listwidget.itemSelectionChanged.connect(self.update_selected_file)
        self.left_layout.addWidget(self.files_listwidget)

        self.buttons_row_layout = QHBoxLayout()
        self.left_layout.addLayout(self.buttons_row_layout)

        self.add_button = QPushButton("+")
        self.add_button.clicked.connect(self.create_rect)
        self.buttons_row_layout.addWidget(self.add_button)

        self.remove_button = QPushButton("-")
        self.buttons_row_layout.addWidget(self.remove_button)
        self.remove_button.clicked.connect(self.remove_rect)

        self.buttons_row_layout.addStretch()

        self.vload_button = QPushButton("V-Load")
        self.buttons_row_layout.addWidget(self.vload_button)

        self.list_dsl_layout = QHBoxLayout()
        self.rect_list_widget = QListWidget()
        self.list_dsl_layout.addWidget(self.rect_list_widget)
        self.rect_list_widget.itemSelectionChanged.connect(self.focus_rect)

        self.dsl_layout = QVBoxLayout()
        self.dsl_layout.setAlignment(Qt.AlignRight)

        self.delete_button = QPushButton("D")
        self.delete_button.setMaximumWidth(30)
        self.delete_button.clicked.connect(self.delete_all_rects)

        self.save_button = QPushButton("S")
        self.save_button.setMaximumWidth(30)
        self.save_button.clicked.connect(self.save_rects)

        self.load_button = QPushButton("L")
        self.load_button.setMaximumWidth(30)
        self.load_button.clicked.connect(self.load_rects)

        self.dsl_layout.addWidget(self.delete_button)
        self.dsl_layout.addWidget(self.save_button)
        self.dsl_layout.addWidget(self.load_button)

        self.list_dsl_layout.addLayout(self.dsl_layout)
        self.left_layout.addLayout(self.list_dsl_layout)

        self.bottom_buttons_row_layout = QHBoxLayout()
        self.left_layout.addLayout(self.bottom_buttons_row_layout)

        self.extract_button = QPushButton("Extract")
        self.extract_button.clicked.connect(
            lambda: self.extract(self.rect_list_widget.itemWidget(self.rect_list_widget.currentItem()))
        )
        self.bottom_buttons_row_layout.addWidget(self.extract_button)

        self.folder_button = QPushButton("F")
        self.folder_button.clicked.connect(self.set_output_folder)
        self.bottom_buttons_row_layout.addWidget(self.folder_button)

        self.extract_all_button = QPushButton("Extract All")
        self.extract_all_button.clicked.connect(self.extract_all)
        self.bottom_buttons_row_layout.addWidget(self.extract_all_button)

        self.output_folder = Path.home() / "Desktop" / "ScreenshotMaker"  # Current user desktop
        self.output_folder_label = QLabel(f"Output folder: {self.output_folder.absolute()}")
        self.left_layout.addWidget(self.output_folder_label)

        self.ocr_layout = QHBoxLayout()
        self.left_layout.addLayout(self.ocr_layout)

        self.run_ocr_button = QPushButton("Run OCR")
        self.run_ocr_button.clicked.connect(self.run_ocr)
        self.ocr_layout.addWidget(self.run_ocr_button)

        self.rect_file_radiobutton = QRadioButton("Rect columns, file rows")
        self.rect_file_radiobutton.setChecked(True)
        self.ocr_layout.addWidget(self.rect_file_radiobutton)

        self.file_rect_radiobutton = QRadioButton("File columns, rect rows")
        self.ocr_layout.addWidget(self.file_rect_radiobutton)

        self.right_layout = QVBoxLayout()
        self.central_layout.addLayout(self.right_layout)

        self.image_displayer = ImageDisplayer(self)
        self.load_pdf_button.clicked.connect(self.load_pdf)
        self.right_layout.addWidget(self.image_displayer)

        self.selected_rect = None

    def run_ocr(self):
        question = QMessageBox.question(
            self,
            "Run OCR",
            f"Running OCR may take a while.\n"
            f"The folder {self.output_folder.absolute()} will be processed.\n Are you sure you want to continue?",
        )
        if question != QMessageBox.Yes:
            return

        self.setCursor(Qt.WaitCursor)
        workbook = Workbook()

        sheet = workbook.active
        if sheet is None:
            sheet = workbook.create_sheet("OCR results")
        else:
            sheet.title = "OCR results"

        if self.rect_file_radiobutton.isChecked():
            header = ["", *[r.name for r in self.output_folder.iterdir().__next__().iterdir()]]
            sheet.append(header)

            for file in self.output_folder.iterdir():
                row = [file.name]
                for rect in file.iterdir():
                    if not rect.is_file():
                        continue
                    text = perform_ocr(rect)
                    row.append(text)
                sheet.append(row)
        else:
            header = ["", *[f.name for f in self.output_folder.iterdir()]]
            sheet.append(header)

            rects = [r for r in self.output_folder.iterdir().__next__().iterdir() if r.is_file()]
            for rect in rects:
                row = [rect.name]
                for file in self.output_folder.iterdir():
                    if not file.is_dir():
                        continue
                    text = perform_ocr(file / rect.name)
                    row.append(text)
                sheet.append(row)

        self.setCursor(Qt.ArrowCursor)

        while True:
            try:
                file, commit = QFileDialog.getSaveFileName(
                    self, "Save file", "../OCR_results.xlsx", "Excel files (*.xlsx)"
                )
                if not commit:
                    return
                workbook.save(file)
                break
            except PermissionError as e:
                QMessageBox.critical(
                    self,
                    "Permission error",
                    f"Please close the file excel file before running OCR again. Exception:\n"
                    f"{e}",
                )

        # Adjust column width
        for idx, col in enumerate(sheet.columns, 1):
            sheet.column_dimensions[get_column_letter(idx)].auto_size = True

        question = QMessageBox.question(self, "OCR successfully finished", "Do you want to open the generated file?")
        if question == QMessageBox.Yes:
            os.startfile(file)

    def focus_rect(self):
        if self.rect_list_widget.currentItem() is None:
            return

        self.image_displayer.go_to_page(self.rect_list_widget.itemWidget(self.rect_list_widget.currentItem()).page)
        self.rect_list_widget.itemWidget(self.rect_list_widget.currentItem()).drawable_rect.is_selected = True
        if self.selected_rect is not None:
            self.selected_rect.is_selected = False
        self.selected_rect = self.rect_list_widget.itemWidget(self.rect_list_widget.currentItem()).drawable_rect
        self.image_displayer.update_scene()

    def add_file(self, file: PDFFile):
        list_widget_item = QListWidgetItem(self.files_listwidget)
        list_widget_item.setText(str(file))
        self.files_listwidget.addItem(list_widget_item)
        self.files_listwidget.setItemWidget(list_widget_item, file)
        self.image_displayer.update_file(file)
        self.files_listwidget.setCurrentItem(list_widget_item)

    def selected_file(self) -> PDFFile:
        return self.files_listwidget.itemWidget(self.files_listwidget.currentItem())

    def delete_all_pdfs(self):
        for i in range(self.files_listwidget.count()):
            self.files_listwidget.takeItem(0)

        self.files_listwidget.clear()
        self.image_displayer.update_file(None)
        self.image_displayer.update_scene()

    def delete_selected_pdf(self):
        self.files_listwidget.takeItem(self.files_listwidget.currentRow())
        self.image_displayer.update_file(None)
        self.image_displayer.update_scene()

    def update_selected_file(self):
        self.image_displayer.update_file(self.selected_file())

    def load_pdf(self):
        file, commit = QFileDialog.getOpenFileName(self, "Open PDF", "", "PDF Files (*.pdf)")
        if not commit:
            return

        self.add_file(PDFFile(file))

    def create_rect(self):
        if self.selected_file() is None:
            QMessageBox.critical(self, "No file selected", "Please load and select a PDF file first")
            return

        rect_item = QListWidgetItem(self.rect_list_widget)

        name = f"{self.selected_file().file_name()}"

        self.rect_list_widget.addItem(rect_item)

        rect = Rect(self, name, self.image_displayer.current_page)
        rect_item.setSizeHint(rect.sizeHint())
        self.rect_list_widget.setItemWidget(rect_item, rect)

        self.image_displayer.update_scene()

    def rename_rect(self):
        item = self.rect_list_widget.currentItem()
        rect = self.rect_list_widget.itemWidget(item)

        new_name, commit = QInputDialog.getText(self, "Rename rect", "New name:", text=rect.name)

        if not commit:
            return

        rect.label.setText(new_name)
        rect.name = new_name

    def remove_rect(self):
        item = self.rect_list_widget.currentItem()
        if item is None:
            return

        self.rect_list_widget.takeItem(self.rect_list_widget.row(item))
        self.image_displayer.update_scene()

    def set_output_folder(self):
        folder = QFileDialog.getExistingDirectory(self, "Select output folder")
        if not folder:
            return

        self.output_folder = Path(folder)
        self.output_folder_label.setText(f"Output folder: {self.output_folder.absolute()}")

    def extract_all(self):
        if self.rect_list_widget.count() == 0:
            QMessageBox.critical(self, "No rects", "Please load some file and create some rects first")
            return

        for i in range(self.rect_list_widget.count()):
            rect = self.rect_list_widget.itemWidget(self.rect_list_widget.item(i))
            self.extract(rect, info=False)

        total = self.rect_list_widget.count() * self.files_listwidget.count()
        QMessageBox.information(self, "Extraction done", f"Successfully extracted {total} rects")

    def extract(self, rect: Rect, info=True):
        if rect is None:
            QMessageBox.critical(self, "No rect selected", "Please select a rect first")
            return

        if not self.output_folder.exists():
            self.output_folder.mkdir(parents=True)

        for file in self.get_files():
            file_folder_path = self.output_folder / str(file)
            file_output_path = file_folder_path / f"{rect.name}.png"
            if not file_folder_path.exists():
                file_folder_path.mkdir(parents=True)

            rect.crop_image(file).save(str(file_output_path.absolute()))

            if info:
                QMessageBox.information(
                    self, "Extraction successful", f"Image extracted to {file_output_path.absolute()}"
                )

    def get_files(self) -> list[PDFFile]:
        return [
            self.files_listwidget.itemWidget(self.files_listwidget.item(i))
            for i in range(self.files_listwidget.count())
        ]

    def delete_all_rects(self):
        if self.rect_list_widget.count() == 0:
            return

        for i in range(self.rect_list_widget.count()):
            self.rect_list_widget.takeItem(0)

        self.rect_list_widget.clear()
        self.image_displayer.update_scene()

    def save_rects(self):
        if self.rect_list_widget.count() == 0:
            QMessageBox.critical(self, "No rects", "There are no rects to save")
            return

        rects = []
        for i in range(self.rect_list_widget.count()):
            rect = self.rect_list_widget.itemWidget(self.rect_list_widget.item(i))
            rects.append(rect.get_pickle())

        path, commit = QFileDialog.getSaveFileName(self, "Save rects", "", "Rects (*.rects)")
        if not commit:
            return

        with open(path, "wb") as f:
            pickle.dump(rects, f)
        QMessageBox.information(self, "Save successful", f"{len(rects)} rects saved successfully")

    def load_rects(self):
        if self.rect_list_widget.count() != 0:
            question = QMessageBox.question(self, "Load rects", "This will delete all the current rects, are you sure?")
            if question == QMessageBox.No:
                return

        path, commit = QFileDialog.getOpenFileName(self, "Select rects file", "", "Rects file (*.rects)")
        if not commit:
            return

        with open(path, "rb") as f:
            l = pickle.load(f)

        self.delete_all_rects()

        for rect in l:
            rect_item = QListWidgetItem(self.rect_list_widget)
            self.rect_list_widget.addItem(rect_item)
            pickle_rect = PickleRect(rect.x, rect.y, rect.width, rect.height, rect.name, rect.page)
            rect = Rect.from_pickle(self, pickle_rect)
            rect_item.setSizeHint(rect.sizeHint())
            self.rect_list_widget.setItemWidget(rect_item, rect)

        if self.image_displayer.file is not None:
            self.image_displayer.update_scene()
